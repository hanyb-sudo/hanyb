#xls      xlsx
# openpyxl  ->  xlsx

from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    #打开文件
    file = load_workbook(filename=path)
    #所有表格的名称
    # print(file.get_sheet_names())
    sheets = file.get_sheet_names()

    #拿出一个表格
    sheet = file.get_sheet_by_name(sheets[0])

    #最大行数
    print(sheet.max_row)
    #最大列数
    print(sheet.max_column)
    #表名
    print(sheet.title)

    # print(sheet.cell(1,1).value)
    #读取一张表的数据

    for lineNum in range(1, sheet.max_row + 1):
        #每一行的数据放在一个列表中
        lineList = []
        for columnNum in range(1, sheet.max_column + 1):
            #拿数据
            value = sheet.cell(lineNum, columnNum).value
            # if value != None:
            lineList.append(value)
        #打印每一行的数据
        print(lineList)

#不能处理xls文件
path = r"C:\Users\Lenovo\Desktop\大数据方法实验\Apriori算法.xlsx"

readXlsxFile(path)