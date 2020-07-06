#xls      xlsx
# openpyxl  ->  xlsx  只能处理xlsx文件不能处理xls文件

from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    dic = {}
    #打开文件
    file = load_workbook(filename=path)
    #所有表格的名称
    # print(file.get_sheet_names())
    sheets = file.get_sheet_names()
    print(sheets)#['Sheet1']
    for sheetName in sheets:
        #一张表所有的数据
        sheetInfo = readData(file, sheetName)
        dic[sheetName] = sheetInfo
    return dic



def readData(file, sheetName):
    # 拿出一个表格
    sheetInfo = []
    sheet = file.get_sheet_by_name(sheetName)
    for lineNum in range(1, sheet.max_row + 1):
        lineList = []
        for columnNum in range(1, sheet.max_column + 1):
            # 拿数据
            value = sheet.cell(row=lineNum, column=columnNum).value
            lineList.append(value)
        sheetInfo.append(lineList)
    return sheetInfo


#不能处理xls文件
path = r"C:\Users\Lenovo\Desktop\课程设计\2015data.xlsx"

dic_Info = readXlsxFile(path)
print(dic_Info.get("Sheet2"))